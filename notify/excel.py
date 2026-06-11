"""
Excel 报告生成：将多类目异动事件导出为 .xlsx 文件。
"""
import logging
import os
from datetime import datetime
from typing import Optional

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl 未安装，Excel 报告功能不可用")


# 事件类型中文映射
_EVENT_LABELS = {
    "NEW_ENTRY": "新进榜",
    "ENTER_TOP10": "冲进TOP10",
    "RANK_UP_50_PLUS_WARNING": "暴升50+",
    "RANK_UP_30_50_WARNING": "急升30-50",
    "RANK_UP_20": "上升20-29",
    "RANK_UP_10": "上升10-19",
    "RANK_UP_5": "上升5-9",
}

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") if HAS_OPENPYXL else None
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11) if HAS_OPENPYXL else None
_WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") if HAS_OPENPYXL else None
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
) if HAS_OPENPYXL else None

_COLUMNS = [
    ("一级类目", 12),
    ("二级类目", 12),
    ("当前排名", 10),
    ("排名变化", 10),
    ("上轮排名", 10),
    ("事件类型", 14),
    ("商品标题", 50),
]


def generate_report(
    events: list[dict],
    output_dir: str,
    timestamp: Optional[datetime] = None,
) -> str:
    """
    生成 Excel 报告文件。

    Parameters
    ----------
    events : list[dict]
        所有事件列表（含 industry_name, category_name 字段）
    output_dir : str
        输出目录
    timestamp : datetime
        报告时间戳，默认 now

    Returns
    -------
    str  生成的文件路径
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl 未安装，无法生成 Excel 报告")
        return ""

    if not events:
        logger.info("无事件数据，跳过 Excel 报告生成")
        return ""

    if timestamp is None:
        timestamp = datetime.now()

    os.makedirs(output_dir, exist_ok=True)
    ts_str = timestamp.strftime("%Y-%m-%d_%H-%M")
    filename = f"榜单异动_{ts_str}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()

    # ── Sheet 1: 汇总 ──────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "汇总"
    _write_header(ws_summary)
    _write_events(ws_summary, sorted(events, key=_sort_key))

    # ── Sheet 2-N: 按一级类目分组 ──────────────────────────────
    l1_groups: dict[str, list[dict]] = {}
    for ev in events:
        l1 = ev.get("industry_name", "") or "未知"
        l1_groups.setdefault(l1, []).append(ev)

    for l1_name in sorted(l1_groups.keys()):
        sheet_name = l1_name[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)
        _write_header(ws)
        l1_events = sorted(l1_groups[l1_name], key=_sort_key)
        _write_events(ws, l1_events)

    wb.save(filepath)
    logger.info("Excel 报告已生成: %s", filepath)
    return filepath


def _sort_key(ev: dict) -> tuple:
    """排序：一级类目 → 二级类目 → 排名变化(降序) → 当前排名。"""
    delta = ev.get("rank_delta") or 0
    return (
        ev.get("industry_name", ""),
        ev.get("category_name", ""),
        -delta,
        ev.get("rank_current") or 999,
    )


def _write_header(ws) -> None:
    """写入表头行。"""
    for col_idx, (name, width) in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width


def _write_events(ws, events: list[dict]) -> None:
    """写入事件数据行。"""
    for row_idx, ev in enumerate(events, 2):
        rank_prev = ev.get("rank_previous")
        rank_delta = ev.get("rank_delta")

        values = [
            ev.get("industry_name", ""),
            ev.get("category_name", ""),
            ev.get("rank_current"),
            f"↑{rank_delta}" if rank_delta else ("新进榜" if rank_prev is None else ""),
            rank_prev if rank_prev is not None else "-",
            _EVENT_LABELS.get(ev.get("event_type", ""), ev.get("event_type", "")),
            ev.get("product_title", ""),
        ]

        is_warning = ev.get("event_type") in (
            "ENTER_TOP10", "RANK_UP_50_PLUS_WARNING", "RANK_UP_30_50_WARNING"
        )

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx <= 6 else "left",
                vertical="center",
                wrap_text=(col_idx == 7),
            )
            if is_warning:
                cell.fill = _WARN_FILL

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    if events:
        last_col = chr(ord("A") + len(_COLUMNS) - 1)
        ws.auto_filter.ref = f"A1:{last_col}{len(events) + 1}"
