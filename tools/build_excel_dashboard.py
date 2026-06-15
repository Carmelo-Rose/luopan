from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SOURCE = Path("data/dashboard_source.xlsx")
OUTPUT = Path("data/dashboard_visualized.xlsx")

NAVY = "132238"
BLUE = "2F75B5"
CYAN = "35A7C4"
ORANGE = "F28E2B"
RED = "E15759"
GREEN = "59A14F"
PURPLE = "8F6BB3"
LIGHT = "F3F6FA"
MID = "D9E2F3"
WHITE = "FFFFFF"
TEXT = "243447"
MUTED = "667085"
THIN = Side(style="thin", color="D6DCE5")


def rank_delta(value) -> int:
    text = str(value or "")
    if text.startswith("↑"):
        try:
            return int(text[1:])
        except ValueError:
            return 0
    return 0


def style_card(ws, cell_range: str, fill: str, title: str, value, note: str):
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":")[0]]
    top_left.fill = PatternFill("solid", fgColor=fill)
    top_left.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    top_left.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_left.value = f"{title}\n{value}\n{note}"
    top_left.font = Font(color=WHITE, bold=True, size=12)


def set_chart_style(chart, title: str, width=12, height=7):
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    chart.legend = None


wb = load_workbook(SOURCE)
if "仪表盘" in wb.sheetnames:
    del wb["仪表盘"]
if "_图表数据" in wb.sheetnames:
    del wb["_图表数据"]

summary = wb["汇总"]
headers = [c.value for c in summary[1]]
rows = [dict(zip(headers, row)) for row in summary.iter_rows(min_row=2, values_only=True)]
coverage = wb["采集覆盖"]
coverage_rows = list(coverage.iter_rows(min_row=2, values_only=True))

l1_counts = Counter(r["一级类目"] for r in rows)
l2_counts = Counter(r["二级类目"] for r in rows)
type_counts = Counter(r["事件类型"] for r in rows)
rank_buckets = Counter(
    "TOP10" if r["当前排名"] <= 10
    else "11-50" if r["当前排名"] <= 50
    else "51-100" if r["当前排名"] <= 100
    else "101-200"
    for r in rows
)
top_movers = sorted(rows, key=lambda r: (-rank_delta(r["排名变化"]), r["当前排名"]))[:10]

data = wb.create_sheet("_图表数据")
data.sheet_state = "hidden"

sections = {
    "l1": (1, [("一级类目", "异动数")] + l1_counts.most_common()),
    "l2": (6, [("二级类目", "异动数")] + l2_counts.most_common(10)),
    "types": (10, [("事件类型", "数量")] + type_counts.most_common()),
    "ranks": (14, [("当前排名区间", "数量")] + [(x, rank_buckets[x]) for x in ("TOP10", "11-50", "51-100", "101-200")]),
}
for _, (start_col, table) in sections.items():
    for row_idx, values in enumerate(table, 1):
        for col_offset, value in enumerate(values):
            data.cell(row_idx, start_col + col_offset, value)

ws = wb.create_sheet("仪表盘", 0)
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A6"
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.tabColor = BLUE

for col in range(1, 17):
    ws.column_dimensions[chr(64 + col)].width = 12
for row in range(1, 55):
    ws.row_dimensions[row].height = 21

ws.merge_cells("A1:P3")
title = ws["A1"]
title.value = "罗盘商品卡榜异动仪表盘\n2026-06-13 11:46"
title.fill = PatternFill("solid", fgColor=NAVY)
title.font = Font(color=WHITE, bold=True, size=22)
title.alignment = Alignment(horizontal="left", vertical="center")

success_count = sum(1 for r in coverage_rows if r[2] in ("有异动", "无异动", "首次基线"))
total_products = sum((r[3] or 0) for r in coverage_rows)
urgent_count = type_counts["冲进TOP10"] + type_counts["暴升50+"] + type_counts["急升30-50"]
top10_count = rank_buckets["TOP10"]

style_card(ws, "A5:D8", BLUE, "本轮异动", f"{len(rows)} 条", "较上一轮排名变化")
style_card(ws, "E5:H8", GREEN, "采集覆盖", f"{success_count} 个二级类目", f"共 {total_products:,} 条商品")
style_card(ws, "I5:L8", ORANGE, "重点机会", f"{urgent_count} 条", "冲进TOP10 / 急升30+")
style_card(ws, "M5:P8", RED, "当前TOP10", f"{top10_count} 条", "需优先跟进")

for cell_range, text in [
    ("A10:H10", "异动主要来自哪些一级类目"),
    ("I10:P10", "事件严重度构成"),
    ("A27:H27", "异动最多的二级类目 TOP10"),
    ("I27:P27", "当前排名区间分布"),
    ("A44:P44", "排名跃升最快的商品 TOP10"),
]:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=MID)
    cell.font = Font(color=TEXT, bold=True, size=12)
    cell.alignment = Alignment(vertical="center")

l1_chart = BarChart()
l1_chart.type = "col"
set_chart_style(l1_chart, "", width=14, height=8)
l1_chart.add_data(Reference(data, min_col=2, min_row=1, max_row=1 + len(l1_counts)), titles_from_data=True)
l1_chart.set_categories(Reference(data, min_col=1, min_row=2, max_row=1 + len(l1_counts)))
l1_chart.y_axis.title = "异动数"
l1_chart.dataLabels = DataLabelList()
l1_chart.dataLabels.showVal = True
l1_chart.dataLabels.showCatName = False
l1_chart.dataLabels.showSerName = False
l1_chart.series[0].graphicalProperties.solidFill = BLUE
ws.add_chart(l1_chart, "A11")

type_chart = DoughnutChart()
type_chart.add_data(Reference(data, min_col=11, min_row=1, max_row=1 + len(type_counts)), titles_from_data=True)
type_chart.set_categories(Reference(data, min_col=10, min_row=2, max_row=1 + len(type_counts)))
type_chart.title = ""
type_chart.height = 8
type_chart.width = 14
type_chart.holeSize = 55
type_chart.dataLabels = DataLabelList()
type_chart.dataLabels.showPercent = True
type_chart.dataLabels.showVal = False
type_chart.dataLabels.showCatName = False
type_chart.dataLabels.showSerName = False
type_chart.dataLabels.showLeaderLines = True
colors = [RED, ORANGE, PURPLE, BLUE, CYAN, GREEN, "A5A5A5"]
type_chart.series[0].data_points = [
    DataPoint(idx=i, spPr=GraphicalProperties(solidFill=color))
    for i, color in enumerate(colors[:len(type_counts)])
]
ws.add_chart(type_chart, "I11")

l2_chart = BarChart()
l2_chart.type = "bar"
set_chart_style(l2_chart, "", width=14, height=8)
l2_chart.add_data(Reference(data, min_col=7, min_row=1, max_row=11), titles_from_data=True)
l2_chart.set_categories(Reference(data, min_col=6, min_row=2, max_row=11))
l2_chart.dataLabels = DataLabelList()
l2_chart.dataLabels.showVal = True
l2_chart.dataLabels.showCatName = False
l2_chart.dataLabels.showSerName = False
l2_chart.series[0].graphicalProperties.solidFill = CYAN
ws.add_chart(l2_chart, "A28")

rank_chart = BarChart()
rank_chart.type = "col"
set_chart_style(rank_chart, "", width=14, height=8)
rank_chart.add_data(Reference(data, min_col=15, min_row=1, max_row=5), titles_from_data=True)
rank_chart.set_categories(Reference(data, min_col=14, min_row=2, max_row=5))
rank_chart.dataLabels = DataLabelList()
rank_chart.dataLabels.showVal = True
rank_chart.dataLabels.showCatName = False
rank_chart.dataLabels.showSerName = False
rank_chart.series[0].graphicalProperties.solidFill = ORANGE
ws.add_chart(rank_chart, "I28")

table_headers = ["一级类目", "二级类目", "当前排名", "排名变化", "事件类型", "商品标题"]
for col, header in enumerate(table_headers, 1):
    cell = ws.cell(45, col, header)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
widths = [14, 22, 11, 11, 14, 70]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + col)].width = width

for row_idx, item in enumerate(top_movers, 46):
    values = [item[h] for h in table_headers]
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row_idx, col_idx, value)
        cell.fill = PatternFill("solid", fgColor=WHITE if row_idx % 2 == 0 else LIGHT)
        cell.font = Font(color=TEXT, size=10)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = Alignment(
            horizontal="left" if col_idx in (1, 2, 5, 6) else "center",
            vertical="center",
            wrap_text=col_idx == 6,
        )
    ws.row_dimensions[row_idx].height = 32

ws.merge_cells("G45:P55")
insight = ws["G45"]
top_l1, top_l1_count = l1_counts.most_common(1)[0]
top_l2, top_l2_count = l2_counts.most_common(1)[0]
insight.value = (
    "本轮观察\n\n"
    f"• {top_l1}贡献 {top_l1_count} 条异动，占本轮 {top_l1_count / len(rows):.0%}。\n"
    f"• 异动最多的二级类目是{top_l2}，共 {top_l2_count} 条。\n"
    f"• 重点机会共 {urgent_count} 条，其中冲进TOP10 {type_counts['冲进TOP10']} 条。\n"
    f"• 当前排名仍在101-200的异动商品有 {rank_buckets['101-200']} 条，建议优先筛选排名跃升30+的商品。"
)
insight.fill = PatternFill("solid", fgColor=LIGHT)
insight.font = Font(color=TEXT, size=12)
insight.alignment = Alignment(vertical="top", wrap_text=True)
insight.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

for existing in wb.worksheets[1:]:
    if existing.title != "_图表数据":
        existing.sheet_view.showGridLines = False

wb.active = 0
wb.save(OUTPUT)
print(OUTPUT.resolve())
