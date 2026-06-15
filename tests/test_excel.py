from openpyxl import load_workbook

from notify.excel import generate_report


def test_report_includes_collection_coverage_without_events(tmp_path):
    category_results = [
        {
            "industry_name": "图书教育",
            "category_name": "文教文化用品",
            "status": "首次基线",
            "products": 200,
            "events": 0,
        },
        {
            "industry_name": "智能家居",
            "category_name": "家具",
            "status": "采集失败",
            "products": 0,
            "events": 0,
        },
    ]

    report_path = generate_report(
        [], str(tmp_path), category_results=category_results,
    )

    workbook = load_workbook(report_path, read_only=True, data_only=True)
    assert workbook.sheetnames == ["汇总", "采集覆盖"]
    coverage = list(workbook["采集覆盖"].iter_rows(values_only=True))
    assert coverage[1] == ("图书教育", "文教文化用品", "首次基线", 200, 0)
    assert coverage[2] == ("智能家居", "家具", "采集失败", 0, 0)
